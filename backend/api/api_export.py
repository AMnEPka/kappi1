"""
API endpoints for exporting data (Excel, etc.).
"""

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import FileResponse
from datetime import date
from typing import List, Optional
import tempfile

from openpyxl import Workbook  # pyright: ignore[reportMissingModuleSource]
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment  # pyright: ignore[reportMissingModuleSource]

from config.config_init import db
from models.models_init import User, Execution
from services.services_init import get_current_user, has_permission, can_access_project
from utils.db_utils import parse_from_mongo
from utils.audit_utils import log_audit

router = APIRouter()


@router.get("/projects/{project_id}/sessions/{session_id}/export-excel")
async def export_session_to_excel(project_id: str, session_id: str, current_user: User = Depends(get_current_user)):
    """Export session execution results to Excel file (requires results_export_all or project access)"""
    
    # Check if user can export all results or has access to project
    if not await has_permission(current_user, 'results_export_all'):
        if not await can_access_project(current_user, project_id):
            raise HTTPException(status_code=403, detail="У вас нет доступа к этому проекту")
    
    # Get project info (single fetch — used for name + system_input_target)
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Проект не найден")

    project_name = project.get('name') or "Неизвестный проект"
    # ОПЭ / ПЭ — задаётся при создании проекта, определяет, какой столбец
    # критичности скрипта использовать в протоколе.
    system_input_target = project.get('system_input_target') or "ОПЭ"
    
    # Get executions for this session
    executions = await db.executions.find(
        {"project_id": project_id, "execution_session_id": session_id},
        {"_id": 0}
    ).sort("executed_at", 1).to_list(1000)
    
    if not executions:
        raise HTTPException(status_code=404, detail="Результаты выполнения не найдены")
    
    # Pre-fetch all scripts and hosts in batch to avoid N+1 queries
    script_ids = list(set(e.get("script_id") for e in executions if e.get("script_id")))
    host_ids = list(set(e.get("host_id") for e in executions if e.get("host_id")))
    
    # Batch fetch scripts
    scripts_docs = await db.scripts.find({"id": {"$in": script_ids}}, {"_id": 0}).to_list(len(script_ids))
    scripts_cache = {s["id"]: s for s in scripts_docs}
    
    # Batch fetch hosts
    hosts_docs = await db.hosts.find({"id": {"$in": host_ids}}, {"_id": 0}).to_list(len(host_ids))
    hosts_cache = {h["id"]: h for h in hosts_docs}
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Протокол испытаний"
    
    # Define styles
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    yellow_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    header_font = Font(bold=True, size=11)
    
    # A1: Протокол
    ws['A1'] = "Протокол проведения испытаний №"
    ws['A1'].font = Font(bold=True, size=14)
    
    # A2: Информационная система
    ws['A2'] = "Информационная система ..."
    ws['A2'].font = Font(bold=True, size=12)
    
    # A4: Место проведения
    ws['A4'] = "Место проведения испытаний: удалённо"
    
    # A5: Дата
    ws['A5'] = f"Дата проведения испытаний: {date.today().strftime('%d.%m.%Y')}"
    
    # Row 8: Table headers
    headers = [
        "№ п/п",
        "Реализация требования ИБ",
        "№ п/п",
        "Описание методики испытания",
        "Критерий успешного прохождения испытания",
        "Результат испытания",
        "Комментарии",
        "Уровень критичности"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = header_font
        cell.fill = yellow_fill
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Значения, записываемые в столбец F ("Результат испытания").
    # Формулы СЧЁТЕСЛИ в сводной таблице M3:S8 ищут точно эти строки,
    # поэтому мы используем "Пройдено" / "Не пройдено" (не "Пройдена"/"Не пройдена").
    # "Частично" и "Не применимо" оператор проставляет вручную после экспорта
    # для статусов "Ошибка" / "Оператор".
    RESULT_PASSED = "Пройдено"
    RESULT_FAILED = "Не пройдено"
    # Поле в Script с уровнем критичности для текущего target (ОПЭ/ПЭ)
    criticality_field = (
        "non_compliance_criticality_ope"
        if system_input_target == "ОПЭ"
        else "non_compliance_criticality_pe"
    )

    def map_check_status(check_status: Optional[str]) -> str:
        if check_status == "Пройдена":
            return RESULT_PASSED
        if check_status == "Не пройдена":
            return RESULT_FAILED
        if check_status == "Оператор":
            return "Требует участия оператора"
        # "Ошибка" и всё прочее — оставляем как есть
        return check_status or "Не определён"

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Data rows starting from row 9
    row_num = 9
    for idx, execution_data in enumerate(executions, 1):
        execution = Execution(**parse_from_mongo(execution_data))

        # Get script and host info from pre-fetched cache (no N+1 queries)
        script = scripts_cache.get(execution.script_id, {})
        host = hosts_cache.get(execution.host_id, {})

        test_methodology = script.get('test_methodology', '') or ''
        success_criteria = script.get('success_criteria', '') or ''

        result = map_check_status(execution.check_status)

        # "Комментарии": сначала описание ошибки (если есть), затем с новой
        # строки — информация о хосте и учётной записи (ранее жила в
        # столбце "Уровень критичности").
        comment_lines: List[str] = []
        if execution.error_code and execution.error_description:
            comment_lines.append(
                f"Код ошибки: {execution.error_code}. {execution.error_description}"
            )
        elif execution.error:
            comment_lines.append(execution.error)

        if host:
            hostname_or_ip = host.get('hostname', 'неизвестно')
            username = host.get('username', 'неизвестно')
            comment_lines.append(
                f"Испытания проводились на хосте {hostname_or_ip} "
                f"под учетной записью {username}"
            )
        comments = "\n".join(comment_lines)

        # Уровень критичности берём из скрипта, выбирая столбец согласно
        # system_input_target проекта (ОПЭ или ПЭ).
        criticality = script.get(criticality_field) or "Нет"

        row_data = [
            idx,                # № п/п
            "",                 # Реализация требования ИБ
            idx,                # № п/п (дубликат)
            test_methodology,   # Описание методики
            success_criteria,   # Критерий успешного прохождения
            result,             # Результат испытания
            comments,           # Комментарии (ошибка + хост/пользователь)
            criticality,        # Уровень критичности проверки (ОПЭ/ПЭ)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

        row_num += 1

    # last_data_row — номер последней строки с данными. Если проверок нет,
    # используем строку 9, чтобы получить валидный (пустой) диапазон F9:F9
    # и не захватить заголовок таблицы.
    last_data_row = row_num - 1 if row_num > 9 else 9

    # ----------------- Сводная таблица (M3:S8) -----------------
    # Блок 1 (M/N/O): распределение результатов испытаний.
    # Блок 2 (Q/R/S): разбивка не пройденных по уровню критичности.
    # "Высокая" в блоке 2 включает и "Высокая (Стоп-фактор)" — поэтому
    # используется СУММ(СЧЁТЕСЛИМН(...)) по обоим значениям.
    f_range = f"F9:F{last_data_row}"
    h_range = f"H9:H{last_data_row}"

    # Ячейки с числами, на которые опираются процентные формулы.
    # N7 = всего; N8 = всего применимых (N7 - "Не применимо").
    # Процент в O считаем от N8 ("всего применимых") — это 100%.
    summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    summary_header_font = Font(bold=True)
    percent_format = "0.0%"

    summary_left = [
        ("M3", "Пройдено",         f'=COUNTIF({f_range},"Пройдено")',     '=N3/N$8'),
        ("M4", "Не пройдено",      f'=COUNTIF({f_range},"Не пройдено")',  '=N4/N$8'),
        ("M5", "Частично",         f'=COUNTIF({f_range},"Частично")',     '=N5/N$8'),
        ("M6", "Не применимо",     f'=COUNTIF({f_range},"Не применимо")', None),
        ("M7", "Всего",            f'=COUNTA({f_range})',                 None),
        ("M8", "Всего применимых", '=N7-N6',                              '=1'),
    ]
    for addr, label, n_formula, o_formula in summary_left:
        row = int(addr[1:])
        label_cell = ws.cell(row=row, column=13, value=label)  # M
        label_cell.font = summary_header_font
        label_cell.fill = summary_fill
        label_cell.border = thin_border
        label_cell.alignment = Alignment(horizontal='left', vertical='center')

        n_cell = ws.cell(row=row, column=14, value=n_formula)  # N
        n_cell.border = thin_border
        n_cell.alignment = Alignment(horizontal='center', vertical='center')

        o_cell = ws.cell(row=row, column=15, value=o_formula)  # O
        o_cell.border = thin_border
        o_cell.alignment = Alignment(horizontal='center', vertical='center')
        if o_formula is not None:
            o_cell.number_format = percent_format

    # Блок справа (Q/R/S) — "Не пройдено" по уровням критичности.
    # "Высокая" объединяет "Высокая" и "Высокая (Стоп-фактор)".
    high_formula = (
        f'=COUNTIFS({f_range},"Не пройдено",{h_range},"Высокая")'
        f'+COUNTIFS({f_range},"Не пройдено",{h_range},"Высокая (Стоп-фактор)")'
    )
    medium_formula = f'=COUNTIFS({f_range},"Не пройдено",{h_range},"Средняя")'
    low_formula = f'=COUNTIFS({f_range},"Не пройдено",{h_range},"Низкая")'

    summary_right = [
        ("Q3", "Высокая", high_formula,   '=R3/N$8'),
        ("Q4", "Средняя", medium_formula, '=R4/N$8'),
        ("Q5", "Низкая",  low_formula,    '=R5/N$8'),
        ("Q6", "Всего",   '=R3+R4+R5',    None),
    ]
    for addr, label, r_formula, s_formula in summary_right:
        row = int(addr[1:])
        label_cell = ws.cell(row=row, column=17, value=label)  # Q
        label_cell.font = summary_header_font
        label_cell.fill = summary_fill
        label_cell.border = thin_border
        label_cell.alignment = Alignment(horizontal='left', vertical='center')

        r_cell = ws.cell(row=row, column=18, value=r_formula)  # R
        r_cell.border = thin_border
        r_cell.alignment = Alignment(horizontal='center', vertical='center')

        s_cell = ws.cell(row=row, column=19, value=s_formula)  # S
        s_cell.border = thin_border
        s_cell.alignment = Alignment(horizontal='center', vertical='center')
        if s_formula is not None:
            s_cell.number_format = percent_format

    # Set column widths
    ws.column_dimensions['A'].width = 8   # № п/п
    ws.column_dimensions['B'].width = 25  # Реализация требования ИБ
    ws.column_dimensions['C'].width = 8   # № п/п
    ws.column_dimensions['D'].width = 35  # Описание методики
    ws.column_dimensions['E'].width = 35  # Критерий успешного прохождения
    ws.column_dimensions['F'].width = 20  # Результат
    ws.column_dimensions['G'].width = 35  # Комментарии (теперь с инфой о хосте)
    ws.column_dimensions['H'].width = 25  # Уровень критичности

    # Сводная таблица справа
    ws.column_dimensions['M'].width = 22
    ws.column_dimensions['N'].width = 10
    ws.column_dimensions['O'].width = 10
    ws.column_dimensions['Q'].width = 14
    ws.column_dimensions['R'].width = 10
    ws.column_dimensions['S'].width = 10

    # Set row heights
    ws.row_dimensions[8].height = 40  # Header row
    
    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        wb.save(tmp_file.name)
        tmp_file_path = tmp_file.name

    # Generate filename
    filename = f"Протокол_испытаний_{date.today().strftime('%d%m%Y')}.xlsx"

    log_audit(
        "26",
        user_id=current_user.id,
        username=current_user.username,
        details={
            "project_name": project_name,
            "project_filename": filename
        }
    )

    return FileResponse(
        path=tmp_file_path,
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

