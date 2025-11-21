"""
services/export.py
Export services for generating reports and Excel files
"""

from io import BytesIO
from datetime import datetime
from typing import List, Optional
from openpyxl import Workbook  # pyright: ignore[reportMissingModuleSource]
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment  # pyright: ignore[reportMissingModuleSource]

from config.config_init import logger, db
from models.models_init import Execution


async def export_executions_to_excel(
    project_id: str,
    session_id: Optional[str] = None
) -> BytesIO:
    """
    Export execution results to Excel file
    """
    try:
        # Fetch executions
        query = {"project_id": project_id}
        if session_id:
            query["execution_session_id"] = session_id
        
        executions = await db.executions.find(query).to_list(None)
        
        # Generate Excel file
        xlsx_bytes = generate_xlsx_file(executions)
        return xlsx_bytes
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        raise


def generate_xlsx_file(executions: List[dict]) -> BytesIO:
    """
    Generate XLSX file from execution results
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Headers
    headers = ["Хост", "Система", "Скрипт", "Статус", "Результат", "Ошибка", "Время"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Data
    for execution in executions:
        status = "✓ OK" if execution.get("success") else "✗ FAILED"
        check_status = execution.get("check_status", "")
        
        row = [
            execution.get("host_id", ""),
            execution.get("system_id", ""),
            execution.get("script_name", ""),
            status,
            execution.get("output", "")[:100],  # First 100 chars
            execution.get("error", "")[:100],   # First 100 chars
            execution.get("executed_at", "")
        ]
        
        ws.append(row)
        
        for cell in ws[ws.max_row]:
            cell.border = border
            if cell.column == 4:  # Status column
                cell.alignment = center_align
                if "OK" in cell.value:
                    cell.font = Font(color="00B050")
            else:
                cell.alignment = left_align
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 20
    
    # Convert to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def create_execution_report(project_id: str) -> dict:
    """
    Create a report summary of project executions
    """
    try:
        executions = await db.executions.find({"project_id": project_id}).to_list(None)
        
        if not executions:
            return {"total": 0, "success": 0, "failed": 0, "success_rate": 0}
        
        total = len(executions)
        success = sum(1 for e in executions if e.get("success"))
        failed = total - success
        success_rate = (success / total * 100) if total > 0 else 0
        
        return {
            "total": total,
            "success": success,
            "failed": failed,
            "success_rate": round(success_rate, 2)
        }
    except Exception as e:
        logger.error(f"Error creating report: {e}")
        raise
