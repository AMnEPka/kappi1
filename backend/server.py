from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Depends, status, Request # pyright: ignore[reportMissingImports]
from fastapi.responses import StreamingResponse, FileResponse # pyright: ignore[reportMissingImports]
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials # pyright: ignore[reportMissingImports]
from starlette.middleware.cors import CORSMiddleware # pyright: ignore[reportMissingImports]
from motor.motor_asyncio import AsyncIOMotorClient # pyright: ignore[reportMissingImports]
import asyncio, logging, json
import os
import uuid
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta, time, date
import contextlib
from openpyxl import Workbook  # pyright: ignore[reportMissingModuleSource]
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment  # pyright: ignore[reportMissingModuleSource]
import tempfile
from typing import Tuple  # pyright: ignore[reportMissingModuleSource]

from config.config_init import *
from models.models_init import *
from services.services_init import *
from utils.db_utils import prepare_for_mongo, parse_from_mongo
from utils.audit_utils import log_audit, _persist_audit_log
from scheduler.scheduler_utils import parse_datetime_param as _parse_datetime_param, calculate_next_run as _calculate_next_run, normalize_run_times as _normalize_run_times, next_daily_occurrence as _next_daily_occurrence
from scheduler.scheduler_worker import scheduler_worker
from api import api_router as auth_api_router

scheduler_task: Optional[asyncio.Task] = None

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix for remaining endpoints
api_router = APIRouter(prefix="/api")

# JWT Configuration
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-secret-key-change-in-production-please-use-strong-random-key')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24

@api_router.get("/permissions", response_model=Dict[str, Any])
async def get_permissions_list():
    """Get all available permissions with descriptions and groups"""
    return {
        "permissions": PERMISSIONS,
        "groups": PERMISSION_GROUPS
    }



# Scheduler API
@api_router.get("/scheduler/jobs", response_model=List[SchedulerJob])
async def list_scheduler_jobs(current_user: User = Depends(get_current_user)):
    query = {} if current_user.is_admin else {"created_by": current_user.id}
    jobs = await db.scheduler_jobs.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    return [SchedulerJob(**parse_from_mongo(job)) for job in jobs]

async def _ensure_project_access(current_user: User, project_id: str):
    if not (await can_access_project(current_user, project_id)):
        raise HTTPException(status_code=403, detail="Access denied to this project")

@api_router.post("/scheduler/jobs", response_model=SchedulerJob)
async def create_scheduler_job(job_input: SchedulerJobCreate, current_user: User = Depends(get_current_user)):
    await require_permission(current_user, 'projects_execute')
    await _ensure_project_access(current_user, job_input.project_id)
    
    # Get project name for logging
    project = await db.projects.find_one({"id": job_input.project_id})
    project_name = project.get('name') if project else "Неизвестный проект"
    
    job = SchedulerJob(
        name=job_input.name,
        project_id=job_input.project_id,
        job_type=job_input.job_type,
        created_by=current_user.id,
    )
    now = datetime.now(timezone.utc)
    if job.job_type == "one_time":
        if not job_input.run_at:
            raise HTTPException(status_code=400, detail="Укажите время запуска")
        job.next_run_at = job_input.run_at if job_input.run_at.tzinfo else job_input.run_at.replace(tzinfo=timezone.utc)
    elif job.job_type == "multi_run":
        if not job_input.run_times:
            raise HTTPException(status_code=400, detail="Укажите список запусков")
        job.run_times = _normalize_run_times(job_input.run_times)
        if not job.run_times:
            raise HTTPException(status_code=400, detail="Нет валидных дат запусков")
        job.remaining_runs = len(job.run_times)
        job.next_run_at = job.run_times[0]
    elif job.job_type == "recurring":
        if not job_input.recurrence_time:
            raise HTTPException(status_code=400, detail="Укажите время повторения")
        start_date = job_input.recurrence_start_date or now.date()
        job.schedule_config = {
            "recurrence_time": job_input.recurrence_time,
            "recurrence_start_date": start_date.isoformat()
        }
        job.next_run_at = _next_daily_occurrence(job.schedule_config, reference=now, initial=True)
    else:
        raise HTTPException(status_code=400, detail="Неверный тип задания")
    doc = prepare_for_mongo(job.model_dump())
    await db.scheduler_jobs.insert_one(doc)
    
    # Логирование создания задания планировщика
    log_audit(
        "29",  # Создание задания планировщика
        user_id=current_user.id,
        username=current_user.username,
        details={
            "job_name": job_input.name,
            "project_name": project_name,
            "job_type_label": _get_job_type_label(job_input.job_type)
        }
    )
    
    return job

def _get_job_type_label(job_type: str) -> str:
    """Get Russian label for job type"""
    labels = {
        "one_time": "Одиночный",
        "multi_run": "Множественный", 
        "recurring": "Ежедневный"
    }
    return labels.get(job_type, job_type)

async def _get_scheduler_job(job_id: str, current_user: User) -> SchedulerJob:
    job_doc = await db.scheduler_jobs.find_one({"id": job_id}, {"_id": 0})
    if not job_doc:
        raise HTTPException(status_code=404, detail="Задание не найдено")
    if not current_user.is_admin and job_doc.get("created_by") != current_user.id:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    return SchedulerJob(**parse_from_mongo(job_doc))

@api_router.put("/scheduler/jobs/{job_id}", response_model=SchedulerJob)
async def update_scheduler_job(job_id: str, job_update: SchedulerJobUpdate, current_user: User = Depends(get_current_user)):
    await require_permission(current_user, 'projects_execute')
    job = await _get_scheduler_job(job_id, current_user)
    
    # Get project name for logging
    project = await db.projects.find_one({"id": job.project_id})
    project_name = project.get('name') if project else "Неизвестный проект"

    # ДЕБАГ: вывести все поля проекта
    print("=== DEBUG PROJECT FIELDS ===")
    print(f"Project ID: {job.project_id}")
    if project:
        print("All project fields:")
        for key, value in project.items():
            print(f"  {key}: {value}")
    else:
        print("Project not found!")
    print("===========================")    
    
    changed = False
    updated_fields = []
    now = datetime.now(timezone.utc)
    
    if job_update.name:
        job.name = job_update.name
        changed = True
        updated_fields.append("название")
    
    if job.job_type == "one_time" and job_update.run_at:
        job.next_run_at = job_update.run_at if job_update.run_at.tzinfo else job_update.run_at.replace(tzinfo=timezone.utc)
        job.status = "active"
        changed = True
        updated_fields.append("время запуска")
    
    if job.job_type == "multi_run" and job_update.run_times:
        job.run_times = _normalize_run_times(job_update.run_times)
        job.remaining_runs = len(job.run_times)
        job.next_run_at = job.run_times[0] if job.run_times else None
        job.status = "active" if job.run_times else "completed"
        changed = True
        updated_fields.append("расписание запусков")
    
    if job.job_type == "recurring":
        recurrence_changed = False
        if job_update.recurrence_time:
            job.schedule_config["recurrence_time"] = job_update.recurrence_time
            recurrence_changed = True
            updated_fields.append("время повторения")
        if job_update.recurrence_start_date:
            job.schedule_config["recurrence_start_date"] = job_update.recurrence_start_date.isoformat()
            recurrence_changed = True
            updated_fields.append("дата начала")
        if recurrence_changed:
            job.next_run_at = _next_daily_occurrence(job.schedule_config, reference=now, initial=True)
            changed = True
    
    if job_update.status in {"active", "paused"}:
        job.status = job_update.status
        status_label = "активен" if job_update.status == "active" else "приостановлен"
        if job.status == "active" and job.job_type == "recurring":
            job.next_run_at = job.next_run_at or _next_daily_occurrence(job.schedule_config, reference=now, initial=True)
        changed = True
        updated_fields.append(f"статус ({status_label})")
    
    job.updated_at = now
    
    if not changed:
        return job
    
    await db.scheduler_jobs.update_one(
        {"id": job.id},
        {"$set": prepare_for_mongo(job.model_dump())}
    )
    
    # Логирование обновления задания планировщика
    log_audit(
        "30",  # Редактирование задания планировщика
        user_id=current_user.id,
        username=current_user.username,
        details={
            "job_name": job.name,
            "project_name": project_name,
            "job_type": job.job_type,
            "job_type_label": _get_job_type_label(job.job_type),
            "updated_fields": updated_fields
        }
    )
    
    return job

@api_router.post("/scheduler/jobs/{job_id}/pause")
async def pause_scheduler_job(job_id: str, current_user: User = Depends(get_current_user)):
    job = await _get_scheduler_job(job_id, current_user)
    
    # Get project name for logging
    project = await db.projects.find_one({"id": job.project_id})
    project_name = project.get('name') if project else "Неизвестный проект"
    
    await db.scheduler_jobs.update_one({"id": job.id}, {"$set": {"status": "paused", "updated_at": datetime.now(timezone.utc).isoformat()}})
    
    # Логирование приостановки задания планировщика
    log_audit(
        "31",  # Задание планировщика приостановлено
        user_id=current_user.id,
        username=current_user.username,
        details={
            "job_name": job.name,
            "project_name": project_name,
            "job_type_label": _get_job_type_label(job.job_type)
        }
    )
    
    return {"message": "Задание приостановлено"}

@api_router.post("/scheduler/jobs/{job_id}/resume")
async def resume_scheduler_job(job_id: str, current_user: User = Depends(get_current_user)):
    job = await _get_scheduler_job(job_id, current_user)
    
    # Get project name for logging
    project = await db.projects.find_one({"id": job.project_id})
    project_name = project.get('name') if project else "Неизвестный проект"
    
    next_run = _calculate_next_run(job, initial=True)
    await db.scheduler_jobs.update_one(
        {"id": job.id},
        {"$set": {"status": "active", "next_run_at": next_run.isoformat() if next_run else None, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Логирование возобновления задания планировщика
    log_audit(
        "32",  # Задание планировщика возобновлено
        user_id=current_user.id,
        username=current_user.username,
        details={
            "job_name": job.name,
            "project_name": project_name,
            "job_type_label": _get_job_type_label(job.job_type)
        }
    )
    
    return {"message": "Задание возобновлено"}

@api_router.delete("/scheduler/jobs/{job_id}")
async def delete_scheduler_job(job_id: str, current_user: User = Depends(get_current_user)):
    job = await _get_scheduler_job(job_id, current_user)
    
    # Get project name for logging
    project = await db.projects.find_one({"id": job.project_id})
    project_name = project.get('name') if project else "Неизвестный проект"
    
    await db.scheduler_jobs.delete_one({"id": job.id})
    await db.scheduler_runs.delete_many({"job_id": job.id})
    
    # Логирование удаления задания планировщика
    log_audit(
        "33",  # Удаление задания планировщика
        user_id=current_user.id,
        username=current_user.username,
        details={
            "job_name": job.name,
            "project_name": project_name,
            "job_type_label": _get_job_type_label(job.job_type)
        }
    )
    
    return {"message": "Задание удалено"}

@api_router.get("/scheduler/jobs/{job_id}/runs", response_model=List[SchedulerRun])
async def get_scheduler_runs(job_id: str, current_user: User = Depends(get_current_user)):
    job = await _get_scheduler_job(job_id, current_user)
    runs = await db.scheduler_runs.find({"job_id": job.id}, {"_id": 0}).sort("started_at", -1).limit(50).to_list(50)
    return [SchedulerRun(**parse_from_mongo(run)) for run in runs]


# Project Execution with Real-time Updates (SSE)
@api_router.get("/projects/{project_id}/execute")
async def execute_project(project_id: str, token: Optional[str] = None, skip_audit_log: bool = False):
    """Execute project with real-time updates via Server-Sent Events (requires projects_execute permission and access to project)"""
    
    # Get current user from token parameter (for SSE which doesn't support headers)
    if not token:
        raise HTTPException(status_code=401, detail="Token required for SSE connection")
    
    current_user = await get_current_user_from_token(token)
    
    # Check permission
    if not await has_permission(current_user, 'projects_execute'):
        raise HTTPException(status_code=403, detail="Permission denied: projects_execute")
    
    # Check project access
    if not await can_access_project(current_user, project_id):
        raise HTTPException(status_code=403, detail="Access denied to this project")
    
    if not skip_audit_log:
        # Получаем имя проекта
        project_doc = await db.projects.find_one({"id": project_id})
        project_name = project_doc.get('name') if project_doc else "Неизвестный проект"
        
    log_audit(
        "23",
        user_id=current_user.id,
        username=current_user.username,
        details={
            "project_name": project_name
        }
    )
    
    async def event_generator():
        try:
            # Store user_id for executions
            user_id = current_user.id
            
            # Get project
            project = await db.projects.find_one({"id": project_id}, {"_id": 0})
            if not project:
                yield f"data: {json.dumps({'type': 'error', 'message': 'Проект не найден'})}\n\n"
                return
            
            # Create unique session ID for this execution
            session_id = str(uuid.uuid4())
            
            # Don't update project status - projects are reusable templates now
            yield f"data: {json.dumps({'type': 'status', 'message': 'Начало выполнения проекта', 'session_id': session_id})}\n\n"
            
            # Get all tasks for this project
            tasks_cursor = db.project_tasks.find({"project_id": project_id}, {"_id": 0})
            tasks = await tasks_cursor.to_list(1000)
            
            if not tasks:
                yield f"data: {json.dumps({'type': 'error', 'message': 'Нет заданий для выполнения'})}\n\n"
                return
            
            total_tasks = len(tasks)
            completed_tasks = 0
            failed_tasks = 0
            
            yield f"data: {json.dumps({'type': 'info', 'message': f'Всего заданий: {total_tasks}'})}\n\n"
            
            # Process each task (each task = one host with multiple scripts)
            for task in tasks:
                task_obj = ProjectTask(**parse_from_mongo(task))
                
                # Get host
                host_doc = await db.hosts.find_one({"id": task_obj.host_id}, {"_id": 0})
                if not host_doc:
                    yield f"data: {json.dumps({'type': 'error', 'message': f'Хост не найден: {task_obj.host_id}'})}\n\n"
                    failed_tasks += 1
                    continue
                
                host = Host(**parse_from_mongo(host_doc))
                
                # Get system
                system_doc = await db.systems.find_one({"id": task_obj.system_id}, {"_id": 0})
                if not system_doc:
                    yield f"data: {json.dumps({'type': 'error', 'message': f'Система не найдена: {task_obj.system_id}'})}\n\n"
                    failed_tasks += 1
                    continue
                
                system = System(**parse_from_mongo(system_doc))
                
                # Get scripts
                scripts_cursor = db.scripts.find({"id": {"$in": task_obj.script_ids}}, {"_id": 0})
                scripts = [Script(**parse_from_mongo(s)) for s in await scripts_cursor.to_list(1000)]
                
                if not scripts:
                    yield f"data: {json.dumps({'type': 'error', 'message': 'Скрипты не найдены для задания'})}\n\n"
                    failed_tasks += 1
                    continue
                
                # Update task status
                await db.project_tasks.update_one(
                    {"id": task_obj.id},
                    {"$set": {"status": "running"}}
                )
                
                yield f"data: {json.dumps({'type': 'task_start', 'host_name': host.name, 'host_address': host.hostname, 'system_name': system.name, 'scripts_count': len(scripts)})}\n\n"
                
                # Perform preliminary checks
                loop = asyncio.get_event_loop()
                
                # 1. Check network access
                network_ok, network_msg = await loop.run_in_executor(None, _check_network_access, host)
                yield f"data: {json.dumps({'type': 'check_network', 'host_name': host.name, 'success': network_ok, 'message': network_msg})}\n\n"
                
                if not network_ok:
                    # Mark all scripts as failed with network error
                    for script in scripts:
                        execution = Execution(
                            project_id=project_id,
                            project_task_id=task_obj.id,
                            execution_session_id=session_id,
                            host_id=host.id,
                            system_id=system.id,
                            script_id=script.id,
                            script_name=script.name,
                            success=False,
                            output="",
                            error=network_msg,
                            check_status="Ошибка",
                            executed_by=user_id
                        )
                        exec_doc = prepare_for_mongo(execution.model_dump())
                        await db.executions.insert_one(exec_doc)
                    
                    await db.project_tasks.update_one(
                        {"id": task_obj.id},
                        {"$set": {"status": "failed"}}
                    )
                    failed_tasks += 1
                    yield f"data: {json.dumps({'type': 'task_error', 'host_name': host.name, 'error': network_msg})}\n\n"
                    continue
                
                # 2. Check login and sudo (combined for SSH to avoid multiple connections)
                if host.connection_type == "winrm":
                    # For WinRM, check login first
                    login_ok, login_msg = await loop.run_in_executor(None, _check_winrm_login, host)
                    yield f"data: {json.dumps({'type': 'check_login', 'host_name': host.name, 'success': login_ok, 'message': login_msg})}\n\n"
                    
                    if not login_ok:
                        # Mark all scripts as failed with login error
                        for script in scripts:
                            execution = Execution(
                                project_id=project_id,
                                project_task_id=task_obj.id,
                                execution_session_id=session_id,
                                host_id=host.id,
                                system_id=system.id,
                                script_id=script.id,
                                script_name=script.name,
                                success=False,
                                output="",
                                error=login_msg,
                                check_status="Ошибка",
                            executed_by=user_id
                            )
                            exec_doc = prepare_for_mongo(execution.model_dump())
                            await db.executions.insert_one(exec_doc)
                        
                        await db.project_tasks.update_one(
                            {"id": task_obj.id},
                            {"$set": {"status": "failed"}}
                        )
                        failed_tasks += 1
                        yield f"data: {json.dumps({'type': 'task_error', 'host_name': host.name, 'error': login_msg})}\n\n"
                        continue
                    
                    # Then check admin access
                    sudo_ok, sudo_msg = await loop.run_in_executor(None, _check_admin_access, host)
                    yield f"data: {json.dumps({'type': 'check_sudo', 'host_name': host.name, 'success': sudo_ok, 'message': sudo_msg})}\n\n"
                else:
                    # For SSH, check both login and sudo in one connection
                    login_ok, login_msg, sudo_ok, sudo_msg = await loop.run_in_executor(None, _check_ssh_login_and_sudo, host)
                    yield f"data: {json.dumps({'type': 'check_login', 'host_name': host.name, 'success': login_ok, 'message': login_msg})}\n\n"
                    
                    if not login_ok:
                        # Mark all scripts as failed with login error
                        for script in scripts:
                            execution = Execution(
                                project_id=project_id,
                                project_task_id=task_obj.id,
                                execution_session_id=session_id,
                                host_id=host.id,
                                system_id=system.id,
                                script_id=script.id,
                                script_name=script.name,
                                success=False,
                                output="",
                                error=login_msg,
                                check_status="Ошибка",
                            executed_by=user_id
                            )
                            exec_doc = prepare_for_mongo(execution.model_dump())
                            await db.executions.insert_one(exec_doc)
                        
                        await db.project_tasks.update_one(
                            {"id": task_obj.id},
                            {"$set": {"status": "failed"}}
                        )
                        failed_tasks += 1
                        yield f"data: {json.dumps({'type': 'task_error', 'host_name': host.name, 'error': login_msg})}\n\n"
                        continue
                    
                    # Send sudo check result
                    yield f"data: {json.dumps({'type': 'check_sudo', 'host_name': host.name, 'success': sudo_ok, 'message': sudo_msg})}\n\n"
                
                if not sudo_ok:
                    # Mark all scripts as failed with sudo error
                    for script in scripts:
                        execution = Execution(
                            project_id=project_id,
                            project_task_id=task_obj.id,
                            execution_session_id=session_id,
                            host_id=host.id,
                            system_id=system.id,
                            script_id=script.id,
                            script_name=script.name,
                            success=False,
                            output="",
                            error=sudo_msg,
                            check_status="Ошибка",
                            executed_by=user_id
                        )
                        exec_doc = prepare_for_mongo(execution.model_dump())
                        await db.executions.insert_one(exec_doc)
                    
                    await db.project_tasks.update_one(
                        {"id": task_obj.id},
                        {"$set": {"status": "failed"}}
                    )
                    failed_tasks += 1
                    yield f"data: {json.dumps({'type': 'task_error', 'host_name': host.name, 'error': sudo_msg})}\n\n"
                    continue
                
                # All checks passed, proceed with script execution
                task_success = True
                task_results = []
                scripts_completed = 0
                
                try:
                    # Execute scripts sequentially on the same host with one connection
                    for idx, script in enumerate(scripts, 1):
                        # Get reference data for this script
                        reference_data = task_obj.reference_data.get(script.id, '') if task_obj.reference_data else ''
                        
                        # Use processor if available
                        result = await execute_check_with_processor(host, script.content, script.processor_script, reference_data)
                        
                        scripts_completed += 1
                        yield f"data: {json.dumps({'type': 'script_progress', 'host_name': host.name, 'completed': scripts_completed, 'total': len(scripts)})}\n\n"
                        
                        # Save execution result with session_id
                        execution = Execution(
                            project_id=project_id,
                            project_task_id=task_obj.id,
                            execution_session_id=session_id,  # NEW: Link to session
                            host_id=host.id,
                            system_id=system.id,
                            script_id=script.id,
                            script_name=script.name,
                            success=result.success,
                            output=result.output,
                            error=result.error,
                            check_status=result.check_status,
                            executed_by=user_id
                        )
                        
                        exec_doc = prepare_for_mongo(execution.model_dump())
                        await db.executions.insert_one(exec_doc)
                        
                        task_results.append(execution)
                        
                        if not result.success:
                            task_success = False
                    
                    # Update task status - host is successful if all preliminary checks passed
                    await db.project_tasks.update_one(
                        {"id": task_obj.id},
                        {"$set": {"status": "completed"}}
                    )
                    
                    # Host completed successfully (passed preliminary checks)
                    completed_tasks += 1
                    yield f"data: {json.dumps({'type': 'task_complete', 'host_name': host.name, 'success': True})}\n\n"
                
                except Exception as e:
                    failed_tasks += 1
                    await db.project_tasks.update_one(
                        {"id": task_obj.id},
                        {"$set": {"status": "failed"}}
                    )
                    yield f"data: {json.dumps({'type': 'task_error', 'host_name': host.name, 'error': str(e)})}\n\n"
            
            # Send completion event (don't update project status - project is reusable)
            # completed_tasks = hosts that passed all preliminary checks
            successful_hosts = completed_tasks
            final_status = "completed" if failed_tasks == 0 else "failed"
            yield f"data: {json.dumps({'type': 'complete', 'status': final_status, 'completed': completed_tasks, 'failed': failed_tasks, 'total': total_tasks, 'successful_hosts': successful_hosts, 'session_id': session_id})}\n\n"
        
        except Exception as e:
            logger.error(f"Error during project execution: {str(e)}")
            yield f"data: {json.dumps({'type': 'error', 'message': f'Ошибка: {str(e)}'})}\n\n"
    
    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )

@api_router.get("/projects/{project_id}/execution-failed")  # ← ИЗМЕНИТЬ POST НА GET
async def log_failed_execution(
    project_id: str, 
    current_user: User = Depends(get_current_user)
):
    """Log failed project execution attempts"""
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Проект не найдена")
    
    project_name = project.get('name', 'Неизвестный проект')
    
    log_audit(
        "34",  # Неуспешный запуск проекта
        user_id=current_user.id,
        username=current_user.username,
        details={
            "project_id": project_id,
            "project_name": project_name,
            "failure_reason": "SSH connection failed"  
        }
    )
    
    return {"message": "Failure logged"}


@api_router.get("/audit/logs", response_model=List[AuditLog])
async def get_audit_logs(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    event_types: Optional[str] = None,
    excluded_event_types: Optional[str] = None,
    limit: int = 200,
    current_user: User = Depends(get_current_user)
):
    """Fetch audit logs with optional filters (admin only)"""
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query: Dict[str, Any] = {}
    created_filter: Dict[str, Any] = {}
    
    start_dt = _parse_datetime_param(start_date) if start_date else None
    end_dt = _parse_datetime_param(end_date, end_of_day=True) if end_date else None
    
    if start_dt:
        created_filter["$gte"] = start_dt.isoformat()
    if end_dt:
        created_filter["$lte"] = end_dt.isoformat()
    if created_filter:
        query["created_at"] = created_filter
    
    # Handle event type filtering
    if event_types:
        events = [event.strip() for event in event_types.split(",") if event.strip()]
        if events:
            query["event"] = {"$in": events}
    elif excluded_event_types:
        # Only apply exclusion if no specific events are selected
        excluded_events = [event.strip() for event in excluded_event_types.split(",") if event.strip()]
        if excluded_events:
            query["event"] = {"$nin": excluded_events}
    
    limit = max(1, min(limit, 500))
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return [AuditLog(**parse_from_mongo(log)) for log in logs]


# Get project execution results
@api_router.get("/projects/{project_id}/executions", response_model=List[Execution])
async def get_project_executions(project_id: str, current_user: User = Depends(get_current_user)):
    """Get all execution results for a project"""
    if not await has_permission(current_user, 'results_view_all'):
        if not await can_access_project(current_user, project_id):
            raise HTTPException(status_code=403, detail="Access denied to this project")
    
    executions = await db.executions.find(
        {"project_id": project_id},
        {"_id": 0}
    ).sort("executed_at", -1).to_list(1000)
    
    log_audit( 
        "25", # просмотр результатов проекта
        user_id=current_user.id,
        username=current_user.username,
        details={
            "project_name": project_id
            }
    )
    
    return [Execution(**parse_from_mongo(execution)) for execution in executions]

# Get execution sessions for project (list of unique session runs)
@api_router.get("/projects/{project_id}/sessions")
async def get_project_sessions(project_id: str, current_user: User = Depends(get_current_user)):
    """Get list of execution sessions for a project (requires results_view_all or project access)"""
    # Check access: user must either view all results or have access to the project
    if not await has_permission(current_user, 'results_view_all'):
        if not await can_access_project(current_user, project_id):
            raise HTTPException(status_code=403, detail="Access denied to this project")

    # Get distinct session IDs with their timestamps and check status counts
    # If check_status is not one of the expected values, count it as error
    pipeline = [
        {"$match": {"project_id": project_id, "execution_session_id": {"$ne": None}}},
        {"$group": {
            "_id": "$execution_session_id",
            "executed_at": {"$first": "$executed_at"},
            "count": {"$sum": 1},
            "passed_count": {
                "$sum": {"$cond": [{"$eq": ["$check_status", "Пройдена"]}, 1, 0]}
            },
            "failed_count": {
                "$sum": {"$cond": [{"$eq": ["$check_status", "Не пройдена"]}, 1, 0]}
            },
            "operator_count": {
                "$sum": {"$cond": [{"$eq": ["$check_status", "Оператор"]}, 1, 0]}
            },
            "explicit_error_count": {
                "$sum": {"$cond": [{"$eq": ["$check_status", "Ошибка"]}, 1, 0]}
            },
            # Count executions with null, empty, or unexpected check_status as errors
            "other_count": {
                "$sum": {
                    "$cond": [
                        {
                            "$and": [
                                {"$ne": ["$check_status", "Пройдена"]},
                                {"$ne": ["$check_status", "Не пройдена"]},
                                {"$ne": ["$check_status", "Оператор"]},
                                {"$ne": ["$check_status", "Ошибка"]}
                            ]
                        },
                        1,
                        0
                    ]
                }
            }
        }},
        {"$sort": {"executed_at": -1}}
    ]
    
    sessions = await db.executions.aggregate(pipeline).to_list(1000)
    
    return [{
        "session_id": s["_id"],
        "executed_at": s["executed_at"],
        "total_checks": s["count"],
        "passed_count": s["passed_count"],
        "failed_count": s["failed_count"],
        # Combine explicit errors and unknown statuses
        "error_count": s["explicit_error_count"] + s["other_count"],
        "operator_count": s["operator_count"]
    } for s in sessions]

# Get executions for specific session
@api_router.get("/projects/{project_id}/sessions/{session_id}/executions", response_model=List[Execution])
async def get_session_executions(project_id: str, session_id: str, current_user: User = Depends(get_current_user)):
    """Get all executions for a specific session (requires results_view_all or project access)"""
    # Check if user can view all results or has access to project
    if not await has_permission(current_user, 'results_view_all'):
        if not await can_access_project(current_user, project_id):
            raise HTTPException(status_code=403, detail="Access denied to this project")
    
    executions = await db.executions.find(
        {"project_id": project_id, "execution_session_id": session_id},
        {"_id": 0}
    ).sort("executed_at", 1).to_list(1000)

        # Получаем имя проекта
    project_doc = await db.projects.find_one({"id": project_id})
    project_name = project_doc.get('name') if project_doc else "Неизвестный проект"    
    
    # log_audit( 
    #     "25", # просмотр результатов проекта
    #     user_id=current_user.id,
    #     username=current_user.username,
    #     details={
    #         "project_name": project_id
    #         }
    # )
    
    return [Execution(**parse_from_mongo(execution)) for execution in executions]


# API Routes - Execution (Legacy single-script execution)
@api_router.post("/execute")
async def execute_script(execute_req: ExecuteRequest, current_user: User = Depends(get_current_user)):
    """Execute script on selected hosts (legacy endpoint)"""
    # Get script
    script_doc = await db.scripts.find_one({"id": execute_req.script_id}, {"_id": 0})
    if not script_doc:
        raise HTTPException(status_code=404, detail="Скрипт не найден")
    
    script = Script(**parse_from_mongo(script_doc))
    
    # Get system for this script
    system_doc = await db.systems.find_one({"id": script.system_id}, {"_id": 0})
    if not system_doc:
        raise HTTPException(status_code=404, detail="Система не найдена")
    
    system = System(**parse_from_mongo(system_doc))
    
    # Get hosts
    hosts_cursor = db.hosts.find({"id": {"$in": execute_req.host_ids}}, {"_id": 0})
    hosts = [Host(**parse_from_mongo(h)) for h in await hosts_cursor.to_list(1000)]
    
    if not hosts:
        raise HTTPException(status_code=404, detail="Хосты не найдены")
    
    # Execute on all hosts concurrently
    tasks = [execute_check_with_processor(host, script.content, script.processor_script) for host in hosts]
    results = await asyncio.gather(*tasks)
    
    # Save execution records (one per host)
    execution_ids = []
    for host, result in zip(hosts, results):
        execution = Execution(
            host_id=host.id,
            system_id=system.id,
            script_id=script.id,
            script_name=script.name,
            success=result.success,
            output=result.output,
            error=result.error,
            check_status=result.check_status,
            executed_by=current_user.id
        )
        
        doc = prepare_for_mongo(execution.model_dump())
        await db.executions.insert_one(doc)
        execution_ids.append(execution.id)
    
    return {"execution_ids": execution_ids, "results": [r.model_dump() for r in results]}

@api_router.get("/executions", response_model=List[Execution])
async def get_executions(current_user: User = Depends(get_current_user)):
    """Get all executions (requires results_view_all or shows own executions)"""
    if await has_permission(current_user, 'results_view_all'):
        executions = await db.executions.find({}, {"_id": 0}).sort("executed_at", -1).to_list(1000)
    else:
        executions = await db.executions.find({"executed_by": current_user.id}, {"_id": 0}).sort("executed_at", -1).to_list(1000)
    
    return [Execution(**parse_from_mongo(execution)) for execution in executions]

@api_router.get("/executions/{execution_id}", response_model=Execution)
async def get_execution(execution_id: str, current_user: User = Depends(get_current_user)):
    """Get execution by ID"""
    execution = await db.executions.find_one({"id": execution_id}, {"_id": 0})
    if not execution:
        raise HTTPException(status_code=404, detail="Выполнение не найдено")
    
    # Check access
    if not await has_permission(current_user, 'results_view_all'):
        if execution.get('executed_by') != current_user.id:
            raise HTTPException(status_code=403, detail="Access denied")
    
    return Execution(**parse_from_mongo(execution))


# Excel Export Endpoint
@api_router.get("/projects/{project_id}/sessions/{session_id}/export-excel")
async def export_session_to_excel(project_id: str, session_id: str, current_user: User = Depends(get_current_user)):
    """Export session execution results to Excel file (requires results_export_all or project access)"""
    
    # Check if user can export all results or has access to project
    if not await has_permission(current_user, 'results_export_all'):
        if not await can_access_project(current_user, project_id):
            raise HTTPException(status_code=403, detail="Access denied to this project")    
    
    # Get project info
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Проект не найден")

            # Получаем имя проекта
    project_doc = await db.projects.find_one({"id": project_id})
    project_name = project_doc.get('name') if project_doc else "Неизвестный проект"    
    
    # Get executions for this session
    executions = await db.executions.find(
        {"project_id": project_id, "execution_session_id": session_id},
        {"_id": 0}
    ).sort("executed_at", 1).to_list(1000)
    
    if not executions:
        raise HTTPException(status_code=404, detail="Результаты выполнения не найдены")
    
    # Get scripts cache for methodology and criteria
    scripts_cache = {}
    hosts_cache = {}
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Протокол испытаний"
    
    # Define styles
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    yellow_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    header_font = Font(bold=True, size=11)
    
    # A1: Протокол
    ws['A1'] = "Протокол проведения испытаний №"
    ws['A1'].font = Font(bold=True, size=14)
    
    # A2: Информационная система
    ws['A2'] = "Информационная система ..."
    ws['A2'].font = Font(bold=True, size=12)
    
    # A4: Место проведения
    ws['A4'] = "Место проведения испытаний: удалённо"
    
    # A5: Дата
    ws['A5'] = f"Дата проведения испытаний: {date.today().strftime('%d.%m.%Y')}"
    
    # Row 8: Table headers
    headers = [
        "№ п/п",
        "Реализация требования ИБ",
        "№ п/п",
        "Описание методики испытания",
        "Критерий успешного прохождения испытания",
        "Результат испытания",
        "Комментарии",
        "Уровень критичности"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = header_font
        cell.fill = yellow_fill
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Data rows starting from row 9
    row_num = 9
    for idx, execution_data in enumerate(executions, 1):
        execution = Execution(**parse_from_mongo(execution_data))
        
        # Get script info (with caching)
        if execution.script_id not in scripts_cache:
            script_doc = await db.scripts.find_one({"id": execution.script_id}, {"_id": 0})
            scripts_cache[execution.script_id] = script_doc
        script = scripts_cache.get(execution.script_id, {})
        
        # Get host info (with caching)
        if execution.host_id not in hosts_cache:
            host_doc = await db.hosts.find_one({"id": execution.host_id}, {"_id": 0})
            hosts_cache[execution.host_id] = host_doc
        host = hosts_cache.get(execution.host_id, {})
        
        # Prepare data
        test_methodology = script.get('test_methodology', '') or ''
        success_criteria = script.get('success_criteria', '') or ''
        
        # Result mapping
        result_map = {
            "Пройдена": "Пройдена",
            "Не пройдена": "Не пройдена",
            "Ошибка": "Ошибка",
            "Оператор": "Требует участия оператора"
        }
        result = result_map.get(execution.check_status, execution.check_status or "Не определён")
        
        # Level of criticality column - host and username info
        host_info = ""
        if host:
            hostname_or_ip = host.get('hostname', 'неизвестно')
            username = host.get('username', 'неизвестно')
            host_info = f"Испытания проводились на хосте {hostname_or_ip} под учетной записью {username}"
        
        # Write row data
        row_data = [
            idx,  # № п/п
            "",   # Реализация требования ИБ (пусто)
            idx,  # № п/п (дубликат)
            test_methodology,  # Описание методики
            success_criteria,  # Критерий успешного прохождения
            result,  # Результат
            "",  # Комментарии (пусто)
            host_info  # Уровень критичности
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        row_num += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 8   # № п/п
    ws.column_dimensions['B'].width = 25  # Реализация требования ИБ
    ws.column_dimensions['C'].width = 8   # № п/п
    ws.column_dimensions['D'].width = 35  # Описание методики
    ws.column_dimensions['E'].width = 35  # Критерий успешного прохождения
    ws.column_dimensions['F'].width = 20  # Результат
    ws.column_dimensions['G'].width = 25  # Комментарии
    ws.column_dimensions['H'].width = 40  # Уровень критичности
    
    # Set row heights
    ws.row_dimensions[8].height = 40  # Header row
    
    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        wb.save(tmp_file.name)
        tmp_file_path = tmp_file.name

    # Generate filename
    filename = f"Протокол_испытаний_{date.today().strftime('%d%m%Y')}.xlsx"

    log_audit(
        "26",
        user_id=current_user.id,
        username=current_user.username,
        details={
            "project_name": project_name, 
            "project_filename": filename}
    )

    return FileResponse(
        path=tmp_file_path,
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# API Routes - User Management
@api_router.get("/users", response_model=List[UserResponse])
async def get_users(current_user: User = Depends(get_current_user)):
    """Get all users (requires users_manage permission)"""
    await require_permission(current_user, 'users_manage', 'users_view')
    
    users = await db.users.find().to_list(length=None)
    return [UserResponse(**user) for user in users]

@api_router.post("/users", response_model=UserResponse)
async def create_user(user_input: UserCreate, current_user: User = Depends(get_current_user)):
    """Create new user (requires users_manage permission)"""
    await require_permission(current_user, 'users_manage')
    
    # Check if username already exists
    existing = await db.users.find_one({"username": user_input.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Create user
    user = User(
        username=user_input.username,
        full_name=user_input.full_name,
        password_hash=hash_password(user_input.password),
        is_admin=user_input.is_admin,
        created_by=current_user.id
    )
    
    doc = prepare_for_mongo(user.model_dump())
    await db.users.insert_one(doc)
    
    # Логирование создания пользователя
    log_audit(
        "3",  # Создание пользователя
        user_id=current_user.id,
        username=current_user.username,
        details={
            "new_user_id": str(user.id),
            "username": user_input.username,
            "target_full_name": user_input.full_name,
            "new_is_admin": user_input.is_admin,
            "created_by": current_user.username
        }
    )
    
    return UserResponse(**user.model_dump())

@api_router.put("/users/{user_id}", response_model=UserResponse)
async def update_user(user_id: str, user_update: UserUpdate, current_user: User = Depends(get_current_user)):
    """Update user (requires users_manage permission)"""
    await require_permission(current_user, 'users_manage')
    
    # Find user
    user_doc = await db.users.find_one({"id": user_id})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    old_user = User(**user_doc)
    
    # Подготовка данных для логирования изменений
    changed_fields = {}
    update_data = {}
    
    for field, new_value in user_update.model_dump().items():
        if new_value is not None:
            old_value = getattr(old_user, field, None)
            if old_value != new_value:
                changed_fields[field] = {
                    "old": old_value,
                    "new": new_value
                }
                update_data[field] = new_value
    
    # Update fields
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    # Логирование обновления пользователя
    if changed_fields:
        log_audit(
            "4",  # Редактирование пользователя
            user_id=current_user.id,
            username=current_user.username,
            details={
                "username": old_user.username,
                "target_full_name": old_user.full_name
            }
        )
    
    # Return updated user
    updated_doc = await db.users.find_one({"id": user_id})
    return UserResponse(**updated_doc)

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    """Delete user and reassign their data to admin (requires users_manage permission)"""
    await require_permission(current_user, 'users_manage')
    
    # Cannot delete yourself
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    # Find user to delete for logging
    user_to_delete_doc = await db.users.find_one({"id": user_id})
    if not user_to_delete_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    user_to_delete = User(**user_to_delete_doc)
    
    # Find admin user
    admin_user = await db.users.find_one({"is_admin": True})
    if not admin_user:
        raise HTTPException(status_code=500, detail="No admin user found")
    admin_id = admin_user['id']
    
    # Подсчет данных для перепривязки (для логирования)
    stats = {
        "categories_reassigned": 0,
        "systems_reassigned": 0,
        "scripts_reassigned": 0,
        "hosts_reassigned": 0,
        "projects_reassigned": 0,
        "executions_reassigned": 0,
        "roles_removed": 0,
        "project_access_removed": 0
    }
    
    # Reassign all data to admin
    categories_result = await db.categories.update_many({"created_by": user_id}, {"$set": {"created_by": admin_id}})
    stats["categories_reassigned"] = categories_result.modified_count
    
    systems_result = await db.systems.update_many({"created_by": user_id}, {"$set": {"created_by": admin_id}})
    stats["systems_reassigned"] = systems_result.modified_count
    
    scripts_result = await db.scripts.update_many({"created_by": user_id}, {"$set": {"created_by": admin_id}})
    stats["scripts_reassigned"] = scripts_result.modified_count
    
    hosts_result = await db.hosts.update_many({"created_by": user_id}, {"$set": {"created_by": admin_id}})
    stats["hosts_reassigned"] = hosts_result.modified_count
    
    projects_result = await db.projects.update_many({"created_by": user_id}, {"$set": {"created_by": admin_id}})
    stats["projects_reassigned"] = projects_result.modified_count
    
    executions_result = await db.executions.update_many({"executed_by": user_id}, {"$set": {"executed_by": admin_id}})
    stats["executions_reassigned"] = executions_result.modified_count
    
    # Delete user roles
    roles_result = await db.user_roles.delete_many({"user_id": user_id})
    stats["roles_removed"] = roles_result.deleted_count
    
    # Delete project access
    access_result = await db.project_access.delete_many({"user_id": user_id})
    stats["project_access_removed"] = access_result.deleted_count
    
    # Delete user
    result = await db.users.delete_one({"id": user_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Логирование удаления пользователя
    log_audit(
        "5",  # Удаление пользователя
        user_id=current_user.id,
        username=current_user.username,
        details={
            "deleted_user_id": user_id,
            "username": user_to_delete.username,
            "full_name": user_to_delete.full_name,
            "is_admin": user_to_delete.is_admin,
            "reassigned_to_admin": admin_user['username'],
            "data_reassignment_stats": stats,
            "total_objects_reassigned": sum([
                stats["categories_reassigned"],
                stats["systems_reassigned"], 
                stats["scripts_reassigned"],
                stats["hosts_reassigned"],
                stats["projects_reassigned"],
                stats["executions_reassigned"]
            ])
        }
    )
    
    return {"message": "User deleted and data reassigned to admin"}

@api_router.put("/users/{user_id}/password")
async def reset_user_password(user_id: str, password_data: PasswordResetRequest, current_user: User = Depends(get_current_user)):
    """Reset user password (requires users_manage permission)"""
    await require_permission(current_user, 'users_manage')
    
    # Find user for logging
    user_doc = await db.users.find_one({"id": user_id})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    user = User(**user_doc)
    
    # Hash new password
    new_hash = hash_password(password_data.new_password)
    
    # Update password
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"password_hash": new_hash}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Логирование сброса пароля
    log_audit(
        "4",  # Редактирование пользователя (или можно создать отдельное событие для сброса пароля)
        user_id=current_user.id,
        username=current_user.username,
        details={
            "target_username": user.username,
            "target_full_name": user.full_name,
            "changed_fields": "Был изменен пароль"
        }
    )
    
    return {"message": "Password updated successfully"}

@api_router.get("/users/{user_id}/roles")
async def get_user_roles(user_id: str, current_user: User = Depends(get_current_user)):
    """Get user's roles (requires users_manage permission)"""
    await require_permission(current_user, 'users_manage')
    
    user_roles = await db.user_roles.find({"user_id": user_id}).to_list(length=None)
    role_ids = [ur['role_id'] for ur in user_roles]
    
    roles = []
    for role_id in role_ids:
        role_doc = await db.roles.find_one({"id": role_id})
        if role_doc:
            roles.append(Role(**role_doc))
    
    return roles

@api_router.put("/users/{user_id}/roles")
async def assign_user_roles(user_id: str, role_ids: List[str], current_user: User = Depends(get_current_user)):
    """Assign roles to user (requires users_manage permission)"""
    await require_permission(current_user, 'users_manage')
    
    # Find user for logging
    user_doc = await db.users.find_one({"id": user_id})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    user = User(**user_doc)
    
    # Get current roles for comparison
    current_roles_docs = await db.user_roles.find({"user_id": user_id}).to_list(None)
    current_role_ids = [doc["role_id"] for doc in current_roles_docs]
    
    # Delete existing roles
    await db.user_roles.delete_many({"user_id": user_id})
    
    # Add new roles
    for role_id in role_ids:
        await db.user_roles.insert_one({
            "user_id": user_id,
            "role_id": role_id
        })
    
    # Логирование обновления ролей пользователя
    log_audit(
        "4",  # Редактирование пользователя
        user_id=current_user.id,
        username=current_user.username,
        details={
            "username": user.username,
            "target_full_name": user.full_name,
            "changed_fields": "Ролевая модель изменена",
            "old_roles_count": len(current_role_ids),
            "new_roles_count": len(role_ids)
        }
    )
    
    return {"message": "Roles assigned successfully"}


# API Routes - Role Management
@api_router.get("/roles", response_model=List[Role])
async def get_roles(current_user: User = Depends(get_current_user)):
    """Get all roles"""
    roles = await db.roles.find().to_list(length=None)
    return [Role(**role) for role in roles]

@api_router.post("/roles", response_model=Role)
async def create_role(role_input: RoleCreate, current_user: User = Depends(get_current_user)):
    """Create new role (requires roles_manage permission)"""
    await require_permission(current_user, 'roles_manage')
    
    # Получаем общее количество доступных прав
    total_permissions_count = len(PERMISSIONS)  # или ваш способ получения общего числа прав
    
    # Создаем роль
    role = Role(
        name=role_input.name,
        description=role_input.description,
        permissions=role_input.permissions
    )
    
    doc = prepare_for_mongo(role.model_dump())
    await db.roles.insert_one(doc)
    
    # Логирование создания роли
    log_audit(
        "6",  # Создание роли
        user_id=current_user.id,
        username=current_user.username,
        details={
            "role_id": str(role.id),
            "role_name": role_input.name,
            "permissions_count": len(role_input.permissions),
            "total_permissions": total_permissions_count,
            "permissions_ratio": f"{len(role_input.permissions)}/{total_permissions_count}",
            "created_by": current_user.username
        }
    )
    
    return Role(**role.model_dump())

@api_router.put("/roles/{role_id}", response_model=Role)
async def update_role(role_id: str, role_update: RoleUpdate, current_user: User = Depends(get_current_user)):
    """Update role (requires roles_manage permission)"""
    await require_permission(current_user, 'roles_manage')
    
    # Находим текущую роль для логирования
    current_role_doc = await db.roles.find_one({"id": role_id})
    if not current_role_doc:
        raise HTTPException(status_code=404, detail="Role not found")
    
    current_role = Role(**current_role_doc)
    
    # Получаем общее количество доступных прав
    total_permissions_count = len(PERMISSIONS)  # или ваш способ получения общего числа прав
    
    # Подготовка данных для логирования изменений
    changed_fields = {}
    update_data = {}
    
    for field, new_value in role_update.model_dump().items():
        if new_value is not None:
            old_value = getattr(current_role, field, None)
            if old_value != new_value:
                changed_fields[field] = {
                    "old": old_value,
                    "new": new_value
                }
                update_data[field] = new_value
    
    # Update fields
    if update_data:
        result = await db.roles.update_one({"id": role_id}, {"$set": update_data})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Role not found")
    
    # Получаем обновленную роль для логирования новых прав
    updated_role_doc = await db.roles.find_one({"id": role_id})
    updated_role = Role(**updated_role_doc)
    
    # Логирование обновления роли
    log_audit(
        "7",  # Редактирование роли
        user_id=current_user.id,
        username=current_user.username,
        details={
            "role_id": role_id,
            "role_name": updated_role.name,
            "old_permissions_count": len(current_role.permissions),
            "new_permissions_count": len(updated_role.permissions),
            "total_permissions": total_permissions_count,
            "old_permissions_ratio": f"{len(current_role.permissions)}/{total_permissions_count}",
            "new_permissions_ratio": f"{len(updated_role.permissions)}/{total_permissions_count}",
            "changed_fields": list(changed_fields.keys()) if changed_fields else "Нет изменений",
            "updated_by": current_user.username
        }
    )
    
    return Role(**updated_role_doc)

@api_router.delete("/roles/{role_id}")
async def delete_role(role_id: str, current_user: User = Depends(get_current_user)):
    """Delete role (requires roles_manage permission)"""
    await require_permission(current_user, 'roles_manage')
    
    # Находим роль для логирования
    role_doc = await db.roles.find_one({"id": role_id})
    if not role_doc:
        raise HTTPException(status_code=404, detail="Role not found")
    
    role = Role(**role_doc)
    
    # Подсчет пользователей с этой ролью
    users_with_role_count = await db.user_roles.count_documents({"role_id": role_id})
    
    # Delete user-role assignments
    await db.user_roles.delete_many({"role_id": role_id})
    
    # Delete role
    result = await db.roles.delete_one({"id": role_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Role not found")
    
    # Логирование удаления роли
    log_audit(
        "8",  # Удаление роли
        user_id=current_user.id,
        username=current_user.username,
        details={
            "role_id": role_id,
            "role_name": role.name,
            "permissions_count": len(role.permissions),
            "affected_users_count": users_with_role_count,
            "deleted_by": current_user.username
        }
    )
    
    return {"message": "Role deleted successfully"}

@api_router.get("/permissions")
async def get_permissions(current_user: User = Depends(get_current_user)):
    """Get all available permissions"""
    return PERMISSIONS

# Include the routers in the main app
app.include_router(auth_api_router)  # Auth routes from api/ package
app.include_router(api_router)  # Remaining routes
# Минимальная рабочая CORS конфигурация
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Просто хардкодим для теста
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/api/health")
def health():
    return {"status": "healthy"}

@app.on_event("startup")
async def startup_db_init():
    """Initialize database on startup if needed"""
    try:
        # Check if admin user exists
        existing_admin = await db.users.find_one({"username": "admin"})
        
        if not existing_admin:
            logger.info("🚀 Initializing database with admin user and roles...")
            
            # Create admin user
            admin_id = str(uuid.uuid4())
            admin_user = {
                "id": admin_id,
                "username": "admin",
                "full_name": "Администратор",
                "password_hash": hash_password("admin123"),
                "is_active": True,
                "is_admin": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": None
            }
            await db.users.insert_one(admin_user)
            logger.info("✅ Created admin user")
            
            # Create default roles
            roles = [
                {
                    "id": str(uuid.uuid4()),
                    "name": "Администратор",
                    "permissions": list(PERMISSIONS.keys()),
                    "description": "Полный доступ ко всем функциям системы",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": admin_id
                },
                {
                    "id": str(uuid.uuid4()),
                    "name": "Исполнитель",
                    "permissions": ['projects_execute'],
                    "description": "Только выполнение назначенных проектов",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": admin_id
                },
                {
                    "id": str(uuid.uuid4()),
                    "name": "Куратор",
                    "permissions": ['results_view_all', 'results_export_all'],
                    "description": "Просмотр и экспорт всех результатов",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": admin_id
                },
                {
                    "id": str(uuid.uuid4()),
                    "name": "Разработчик проверок",
                    "permissions": [
                        'checks_create', 'checks_edit_own', 'checks_delete_own',
                        'hosts_create', 'hosts_edit_own', 'hosts_delete_own'
                    ],
                    "description": "Создание и редактирование своих проверок и хостов",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": admin_id
                },
                {
                    "id": str(uuid.uuid4()),
                    "name": "Менеджер проектов",
                    "permissions": [
                        'projects_create', 'projects_execute',
                        'results_view_all', 'results_export_all'
                    ],
                    "description": "Создание и выполнение проектов, просмотр результатов",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": admin_id
                }
            ]
            await db.roles.insert_many(roles)
            logger.info(f"✅ Created {len(roles)} default roles")
            
            # Migrate existing data - assign to admin
            categories_updated = await db.categories.update_many(
                {"created_by": {"$exists": False}},
                {"$set": {"created_by": admin_id}}
            )
            systems_updated = await db.systems.update_many(
                {"created_by": {"$exists": False}},
                {"$set": {"created_by": admin_id}}
            )
            scripts_updated = await db.scripts.update_many(
                {"created_by": {"$exists": False}},
                {"$set": {"created_by": admin_id}}
            )
            hosts_updated = await db.hosts.update_many(
                {"created_by": {"$exists": False}},
                {"$set": {"created_by": admin_id}}
            )
            projects_updated = await db.projects.update_many(
                {"created_by": {"$exists": False}},
                {"$set": {"created_by": admin_id}}
            )
            executions_updated = await db.executions.update_many(
                {"executed_by": {"$exists": False}},
                {"$set": {"executed_by": admin_id}}
            )
            
            logger.info(f"✅ Migrated existing data: {categories_updated.modified_count} categories, "
                       f"{systems_updated.modified_count} systems, {scripts_updated.modified_count} scripts, "
                       f"{hosts_updated.modified_count} hosts, {projects_updated.modified_count} projects, "
                       f"{executions_updated.modified_count} executions")
            
            logger.info("✨ Database initialization complete!")
        else:
            logger.info("✅ Database already initialized")
        global scheduler_task
        if scheduler_task is None:
            scheduler_task = asyncio.create_task(scheduler_worker())
    except Exception as e:
        logger.error(f"❌ Database initialization failed: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
    global scheduler_task
    if scheduler_task:
        scheduler_task.cancel()
        with contextlib.suppress(asyncio.CancelledError):
            await scheduler_task
        scheduler_task = None